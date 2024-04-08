# features/utils.py
import hashlib
import json
import os
import shutil
import time
from datetime import datetime
from typing import Dict, Tuple, Any, List, Optional

import pandas as pd
import requests
from openpyxl import load_workbook
from tenacity import retry, stop_after_attempt

from scripts.config import ConfigManager

config_manager = ConfigManager(None)


def md5(string_data: str) -> str:
    """ 计算输入字符串的MD5值。
    Args:
      string_data: 要计算MD5值的输入字符串。

    Returns:
      输入字符串的MD5值。
    """
    string = str(string_data).encode("utf8")
    str_md5 = hashlib.md5(string).hexdigest()
    return str(str_md5)


def get_timestamp() -> str:
    """ 计算当前时间的timestamp值
    Returns:
      当前时间的timestamp值
    """
    t = time.time()
    timestamp = int(round(t * 1000))
    return str(timestamp)


def get_sign(secret_key, data, timestamp, check_data):
    # 重要的事情说三遍：如果验证Data，数据字段必须跟服务端一致。
    if data is None:
        string_data = timestamp + secret_key
    else:
        str_data = json.dumps(data, separators=(',', ':'), ensure_ascii=False, sort_keys=True)
        string_data = timestamp + secret_key + str_data if check_data else timestamp + secret_key
    sign = md5(string_data)
    return sign


def get_headers():
    return {"Content-Type": "application/json", "Charset": "UTF-8"}


def read_excel_tuple_list(excel_file, min_row):
    if excel_file is None or not excel_file:
        return []
    if os.path.exists(excel_file):
        # 判断路径是文件还是目录
        if os.path.isfile(excel_file):
            # 读取 Excel 文件
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            # 将每一行的单元格值转换为元组
            excel_data = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row)]
            return excel_data
    return []


def create_subdirectory(base_directory):
    # 获取当前时间
    current_time = datetime.now()
    # 格式化时间为字符串，作为子目录名
    subdirectory_name = current_time.strftime("%Y-%m-%d")
    # 拼接子目录的完整路径
    subdirectory_path = os.path.join(base_directory, subdirectory_name)
    # 创建子目录
    os.makedirs(subdirectory_path, exist_ok=True)
    return subdirectory_path, subdirectory_name


def generate_file_name(base_directory):
    subdirectory_path, subdirectory_name = create_subdirectory(base_directory)
    # 获取当前时间戳，用来命名新生成文件名
    date_time = int(time.mktime(time.localtime(time.time())))
    file_name = os.path.join(subdirectory_path, str(date_time) + ".xlsx")
    return file_name


def generate_file_path(base_directory, file_name):
    subdirectory_path, subdirectory_name = create_subdirectory(base_directory)
    file_path = os.path.join(subdirectory_path, file_name + ".xlsx")
    return file_path


def copy_and_rename_file(src_path, dest_path):
    # 检查源文件是否存在
    if not os.path.isfile(src_path):
        print(f"错误: 源文件 '{src_path}' 不存在.")
        return

    # 复制并重命名文件
    try:
        shutil.copy(src_path, dest_path)
        print(f"文件已成功复制到 '{dest_path}'.")
    except Exception as e:
        print(f"复制文件时出现错误: {e}")
    return dest_path


def replace_fullwidth_with_halfwidth(text):
    # 构建全角到半角的映射表
    fullwidth_chars = "，。！？（）［］《》：；－"
    halfwidth_chars = ",.!?()[]<>:;-"

    # 确保两个字符串的长度相等
    if len(fullwidth_chars) != len(halfwidth_chars):
        raise ValueError("两个参数的长度必须相等")

    translation_table = str.maketrans(fullwidth_chars, halfwidth_chars)

    # 使用 translate 进行替换
    result = text.translate(translation_table)
    return result


def rename(file_name):
    if os.path.exists(file_name):
        # 获取当前时间戳
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        # 拆分文件名和扩展名
        base_name, extension = os.path.splitext(file_name)
        # 构建新的文件名，添加时间戳
        new_file_path = f"{base_name}_{timestamp}{extension}"
        # 重命名文件
        os.rename(file_name, new_file_path)


def format_datetime_f1(input_datetime_str):
    # 将字符串解析为 datetime 对象
    input_datetime = datetime.strptime(input_datetime_str.strip(), "%Y-%m-%d %H:%M:%S")
    # 格式化为 "2023-12-18 11:26:31.0"
    formatted_datetime_str = input_datetime.strftime("%Y-%m-%d %H:%M:%S.%f")[:21]
    return formatted_datetime_str


@retry(stop=stop_after_attempt(3))
def requests_post(url, data):
    timestamp = get_timestamp()
    sign = get_sign(config_manager.get_secret_key(), data, timestamp, False)
    try:
        post_data = {"appId": config_manager.get_app_id(),
                     "timestamp": timestamp,
                     "sign": sign,
                     "data": data,
                     "ignorePaperboardProcessFlag": config_manager.get_ignore_paperboard_process_flag(),
                     "ignoreEmptyWorkFlowNoFlag": config_manager.get_ignore_carton_process_flag()}
        return requests.post(url=url, headers=get_headers(), json=post_data, verify=False)
    except Exception as err:
        raise err


def compare_times(time_str1, time_str2):
    if time_str1 is None or not time_str1:
        return False
    if time_str2 is None or not time_str2:
        return False
    # 将时间字符串转换为datetime对象
    time1 = datetime.strptime(time_str1.strip(), "%Y-%m-%d %H:%M:%S")
    time2 = datetime.strptime(time_str2.strip(), "%Y-%m-%d %H:%M:%S")

    # 返回比较结果的布尔值
    return time1 < time2


def get_machine_setting_dicts(file_path: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    # 读取 Excel 文件
    df = pd.read_excel(file_path)

    # 使用字典推导式创建两个字典
    machine_line_dict = {row['机床']: row['线别'] for _, row in df.iterrows()}
    machine_process_dict = {row['机床']: row['工序类型'] for _, row in df.iterrows()}

    return machine_line_dict, machine_process_dict


def find_keys_by_value(_dict: Dict[str, Any], value: Any) -> List[str]:
    """
    Find keys by value in a dictionary.

    Args:
        _dict (dict): The input dictionary.
        value (Any): The value to search for.

    Returns:
        List[str]: A list of keys corresponding to the given value.
    """
    keys = [key for key, val in _dict.items() if val == value]
    return keys


def find_first_key_by_value(_dict: Dict[str, Any], value: Any) -> Optional[str]:
    """
    Find the first key by value in a dictionary.

    Args:
        _dict (dict): The input dictionary.
        value (Any): The value to search for.

    Returns:
        Optional[str]: The first key corresponding to the given value, or None if no match is found.
    """
    for key, val in _dict.items():
        if val == value:
            return key
    return None


def format_machine_name(machine_names: List[str]) -> List[str]:
    """
    Format machine names by removing "--空开" part and remove duplicates.

    Args:
        machine_names (List[str]): List of machine names.

    Returns:
        List[str]: List of unique formatted machine names.
    """
    formatted_names = {name.split('--')[0].strip() for name in machine_names}
    return list(formatted_names)
